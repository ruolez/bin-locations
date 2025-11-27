from flask import Flask, render_template, jsonify, request, session, redirect, url_for, send_file
from flask_session import Session
from functools import wraps
from app.database import SQLiteManager, MSSQLManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO
import traceback
import os

app = Flask(__name__)

# Session configuration
app.config['SECRET_KEY'] = os.urandom(24)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = '/app/data/flask_session'
app.config['SESSION_PERMANENT'] = False
# Cookie settings for cross-origin iframe embedding
# Note: SameSite=None requires Secure=True (HTTPS). For HTTP-only setups,
# cross-origin iframe sessions may not work in modern browsers.
app.config['SESSION_COOKIE_SAMESITE'] = 'None'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SECURE_COOKIES', 'true').lower() == 'true'
Session(app)

# Initialize database managers
sqlite_manager = SQLiteManager()
mssql_manager = MSSQLManager(sqlite_manager)


# ============================================================================
# Authentication Decorator
# ============================================================================

def login_required(f):
    """Decorator to require login for protected routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            if request.is_json:
                return jsonify({'success': False, 'message': 'Authentication required', 'auth_required': True}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


@app.after_request
def add_no_cache_headers(response):
    """Add no-cache headers to all responses"""
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# ============================================================================
# Page Routes
# ============================================================================

@app.route('/login')
def login():
    """Login page"""
    return render_template('login.html')


@app.route('/')
@login_required
def index():
    """Main page - Bin Locations Management"""
    return render_template('index.html', username=session.get('username'))


@app.route('/history')
@login_required
def history():
    """History page - Operation History"""
    return render_template('history.html', username=session.get('username'))


@app.route('/settings')
def settings():
    """Settings page - Database Configuration (accessible without login for first-time setup)"""
    return render_template('settings.html', username=session.get('username'))


# ============================================================================
# Configuration API
# ============================================================================

@app.route('/api/config', methods=['GET'])
def get_config():
    """Get MSSQL configuration (without password)"""
    try:
        config = sqlite_manager.get_config()
        if config:
            # Don't return password
            return jsonify({
                'success': True,
                'config': {
                    'server': config['server'],
                    'port': config['port'],
                    'database': config['database'],
                    'username': config['username']
                }
            })
        return jsonify({'success': True, 'config': None})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/config', methods=['POST'])
def save_config():
    """Save MSSQL configuration"""
    try:
        data = request.json
        sqlite_manager.save_config(data)
        return jsonify({'success': True, 'message': 'Configuration saved successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/config/test', methods=['POST'])
def test_connection():
    """Test MSSQL connection"""
    try:
        # Temporarily save config for testing
        data = request.json
        sqlite_manager.save_config(data)

        # Test connection
        result = mssql_manager.test_connection()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# Authentication API
# ============================================================================

@app.route('/api/login', methods=['POST'])
def api_login():
    """Authenticate user"""
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')

        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password required'}), 400

        # Check if database is configured
        config = sqlite_manager.get_config()

        # First-time setup: Allow admin/admin when no database connection configured
        if not config or not config.get('server'):
            if username == 'admin' and password == 'admin':
                session['username'] = 'admin'
                session['auto_id'] = 0
                session['employee_id'] = 0
                return jsonify({
                    'success': True,
                    'message': 'First-time setup login successful. Please configure database connection.',
                    'username': 'admin',
                    'first_time_setup': True,
                    'redirect': '/settings'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'First-time setup: Use admin/admin to configure database connection'
                }), 401

        # Normal authentication via database
        user = mssql_manager.verify_user_credentials(username, password)

        if user:
            session['username'] = user['username']
            session['auto_id'] = user['auto_id']
            session['employee_id'] = user['employee_id']
            return jsonify({'success': True, 'message': 'Login successful', 'username': user['username']})
        else:
            return jsonify({'success': False, 'message': 'Invalid username or password'}), 401

    except Exception as e:
        error_msg = str(e)
        if 'configuration not found' in error_msg.lower():
            return jsonify({
                'success': False,
                'message': 'Database not configured. Please use admin/admin to access settings.',
                'needs_config': True
            }), 400
        return jsonify({'success': False, 'message': error_msg}), 500


@app.route('/api/logout', methods=['POST'])
def api_logout():
    """Logout user"""
    session.clear()
    return jsonify({'success': True, 'message': 'Logged out successfully'})


@app.route('/api/current-user', methods=['GET'])
@login_required
def get_current_user():
    """Get current logged-in user"""
    return jsonify({
        'success': True,
        'username': session.get('username'),
        'auto_id': session.get('auto_id'),
        'employee_id': session.get('employee_id')
    })


# ============================================================================
# Bin Locations API
# ============================================================================

@app.route('/api/bin-locations', methods=['GET'])
@login_required
def get_bin_locations():
    """Get all bin location records"""
    try:
        records = mssql_manager.get_bin_locations()
        return jsonify({'success': True, 'data': records})
    except Exception as e:
        error_msg = str(e)
        if 'configuration not found' in error_msg.lower():
            return jsonify({
                'success': False,
                'message': 'Database not configured. Please go to Settings to configure the connection.',
                'needs_config': True
            }), 400
        return jsonify({'success': False, 'message': error_msg}), 500


@app.route('/api/bin-locations', methods=['POST'])
@login_required
def create_bin_location():
    """Create new bin location record"""
    try:
        data = request.json

        # Validate required fields
        if not data.get('product_upc'):
            return jsonify({'success': False, 'message': 'Product is required'}), 400
        if not data.get('bin_location_id'):
            return jsonify({'success': False, 'message': 'Bin location is required'}), 400

        result = mssql_manager.create_bin_location(data, session['username'])
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/bin-locations/<int:record_id>', methods=['PUT'])
@login_required
def update_bin_location(record_id):
    """Update bin location record"""
    try:
        data = request.json

        # Validate required fields
        if not data.get('product_upc'):
            return jsonify({'success': False, 'message': 'Product is required'}), 400
        if not data.get('bin_location_id'):
            return jsonify({'success': False, 'message': 'Bin location is required'}), 400

        result = mssql_manager.update_bin_location(record_id, data, session['username'])
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/bin-locations/<int:record_id>/adjust', methods=['PATCH'])
@login_required
def adjust_quantity(record_id):
    """Adjust case quantity"""
    try:
        data = request.json
        adjustment = data.get('adjustment', 0)
        notes = data.get('notes')

        if adjustment == 0:
            return jsonify({'success': False, 'message': 'Adjustment cannot be zero'}), 400

        result = mssql_manager.adjust_quantity(record_id, adjustment, session['username'], notes)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/bin-locations/<int:record_id>', methods=['DELETE'])
@login_required
def delete_bin_location(record_id):
    """Delete bin location record"""
    try:
        result = mssql_manager.delete_bin_location(record_id, session['username'])
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/export-excel', methods=['POST'])
@login_required
def export_to_excel():
    """Export filtered bin location records to Excel"""
    try:
        data = request.json
        records = data.get('records', [])

        if not records:
            return jsonify({'success': False, 'message': 'No records to export'}), 400

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Bin Locations"

        # Define column headers
        headers = [
            'Bin Location',
            'Product Name',
            'Product UPC',
            'Case Quantity',
            'Qty per Case',
            'Total Quantity',
            'Bin Location ID',
            'Created At',
            'Last Updated'
        ]

        # Style definitions - Material Design 3 Slate theme
        header_fill = PatternFill(start_color='546e7a', end_color='546e7a', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        alt_row_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
        totals_fill = PatternFill(start_color='eceff1', end_color='eceff1', fill_type='solid')
        totals_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin', color='dadce0'),
            right=Side(style='thin', color='dadce0'),
            top=Side(style='thin', color='dadce0'),
            bottom=Side(style='thin', color='dadce0')
        )

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Track totals
        total_cases = 0
        total_items = 0

        # Write data rows
        for row_num, record in enumerate(records, 2):
            # Extract values with null handling
            bin_location = record.get('BinLocation') or 'N/A'
            product_name = record.get('ProductDescription') or 'N/A'
            product_upc = record.get('ProductUPC') or 'N/A'
            qty_cases = record.get('Qty_Cases') or 0
            unit_qty2 = record.get('UnitQty2')
            total_qty = record.get('TotalQuantity')
            bin_location_id = record.get('BinLocationID') or 'N/A'
            created_at = record.get('CreatedAt') or 'N/A'
            last_update = record.get('LastUpdate') or 'N/A'

            # Parse timestamps to display format
            if created_at != 'N/A':
                try:
                    created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    created_at = created_at.strftime('%m/%d/%Y %I:%M %p')
                except:
                    pass

            if last_update != 'N/A':
                try:
                    last_update = datetime.fromisoformat(last_update.replace('Z', '+00:00'))
                    last_update = last_update.strftime('%m/%d/%Y %I:%M %p')
                except:
                    pass

            # Handle null UnitQty2
            if unit_qty2 is None or unit_qty2 == 0:
                unit_qty2_display = 'Not Set'
                total_qty_display = '—'
            else:
                unit_qty2_display = unit_qty2
                if total_qty is not None:
                    total_qty_display = total_qty
                    total_items += total_qty
                else:
                    total_qty_display = '—'

            total_cases += qty_cases

            # Write row data
            row_data = [
                bin_location,
                product_name,
                product_upc,
                qty_cases,
                unit_qty2_display,
                total_qty_display,
                bin_location_id,
                created_at,
                last_update
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border

                # Right-align numbers
                if col_num in [4, 5, 6]:
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                        # Add thousand separators for numbers
                        if col_num in [4, 6]:
                            cell.number_format = '#,##0'
                    else:
                        cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')

                # Alternating row colors
                if row_num % 2 == 0:
                    cell.fill = alt_row_fill

        # Add totals row
        totals_row = len(records) + 2
        totals_data = [
            'TOTALS',
            '',
            '',
            total_cases,
            '',
            total_items if total_items > 0 else '—',
            f'{len(records)} records',
            '',
            ''
        ]

        for col_num, value in enumerate(totals_data, 1):
            cell = ws.cell(row=totals_row, column=col_num, value=value)
            cell.fill = totals_fill
            cell.font = totals_font
            cell.border = border

            if col_num in [4, 6]:
                cell.alignment = Alignment(horizontal='right')
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal='left')

        # Auto-fit column widths
        column_widths = {
            'A': 18,  # Bin Location
            'B': 35,  # Product Name
            'C': 15,  # Product UPC
            'D': 15,  # Case Quantity
            'E': 14,  # Qty per Case
            'F': 15,  # Total Quantity
            'G': 16,  # Bin Location ID
            'H': 20,  # Created At
            'I': 20   # Last Updated
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Generate filename with Central Time timestamp
        central_time = datetime.now(ZoneInfo("America/Chicago"))
        timestamp = central_time.strftime('%Y%m%d_%H%M%S')
        filename = f'bin_locations_export_{timestamp}.xlsx'

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# Lookup API
# ============================================================================

@app.route('/api/products/search', methods=['GET'])
@login_required
def search_products():
    """Search products by description, UPC, or SKU"""
    try:
        query = request.args.get('q', '')
        search_field = request.args.get('field', 'description')

        # Validate search field parameter
        if search_field not in ['description', 'upc', 'sku']:
            return jsonify({'success': False, 'message': 'Invalid search field'}), 400

        # Allow single % wildcard, otherwise require 2+ characters
        if len(query) < 2 and query != '%':
            return jsonify({'success': True, 'data': []})

        products = mssql_manager.search_products(query, search_field)
        return jsonify({'success': True, 'data': products})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/bins/search', methods=['GET'])
@login_required
def search_bins():
    """Search bin locations"""
    try:
        query = request.args.get('q', '')
        # Allow all queries with at least 1 character (including % wildcard)
        if len(query) < 1:
            return jsonify({'success': True, 'data': []})

        bins = mssql_manager.search_bin_locations(query)
        return jsonify({'success': True, 'data': bins})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/bin-locations/unused', methods=['GET'])
@login_required
def get_unused_bins():
    """Get bin locations that are not used in Items_BinLocations"""
    try:
        unused_bins = mssql_manager.get_unused_bin_locations()
        return jsonify({'success': True, 'data': unused_bins})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# History API
# ============================================================================

@app.route('/api/history', methods=['GET'])
@login_required
def get_history():
    """Get history records with optional filtering"""
    try:
        record_id = request.args.get('record_id', type=int)
        operation_type = request.args.get('operation_type')
        username = request.args.get('username')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        limit = request.args.get('limit', 500, type=int)

        records = mssql_manager.get_history_records(
            record_id=record_id,
            operation_type=operation_type,
            username=username,
            start_date=start_date,
            end_date=end_date,
            limit=limit
        )

        return jsonify({'success': True, 'data': records})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/history/stats', methods=['GET'])
@login_required
def get_history_stats():
    """Get history statistics"""
    try:
        stats = mssql_manager.get_history_stats()
        return jsonify({'success': True, 'data': stats})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# Health Check
# ============================================================================

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'ok'})


# ============================================================================
# Error Handlers
# ============================================================================

@app.errorhandler(404)
def not_found(e):
    """Handle 404 errors"""
    return jsonify({'success': False, 'message': 'Resource not found'}), 404


@app.errorhandler(500)
def server_error(e):
    """Handle 500 errors"""
    return jsonify({'success': False, 'message': 'Internal server error'}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
